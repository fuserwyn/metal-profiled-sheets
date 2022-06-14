import openpyxl
import pandas as pd
import xlsxwriter
import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border

data = pd.read_excel('Профиль.xlsm')
df = pd.DataFrame(data)
workbook = xlsxwriter.Workbook('formula.xlsx')
worksheet = workbook.add_worksheet("Продукция")

def df_to_list(df):
    lst = []
    for i in range (2, len(df.index)):
        lst.append(df.values[i])
    listoflists = []
    for j in range (0, df.count()[0]):
        proflist_data = []
        for k in range (0, len(data.columns)):
            proflist_data.append(lst[j][k])
        listoflists.append(proflist_data)
    return listoflists

def writer():
    lst = []
    lst = df_to_list(df)
    col = 0
    for j in range (len(lst[0])):
        row = 3
        for i in range (0, len(lst)):
            worksheet.set_column(0, 0, 30)
            worksheet.write(row, col + j, lst[i][j])
            row += 1
    row = 0
    workbook.close()

def decor():
    book = openpyxl.load_workbook('formula.xlsx')
    sheet:worksheet = book.worksheets[0]
    sheet["C1"].value = "Справочные величины на 1 м ширины"
    sheet["C2"].value = "при сжатых узких полках"
    sheet["F2"].value = "при сжатых широких полках"
    sheet["A1"].value = "Обозначение профилей"
    sheet["B1"].value = "t, мм"
    sheet["C3"].value = "Ix, cм4"
    sheet["D3"].value = "Wx1, cм3"
    sheet["E3"].value = "Wx2, cм3"
    sheet["F3"].value = "Ix, cм4"
    sheet["G3"].value = "Wx1, cм3"
    sheet["H3"].value = "Wx2, cм3"
    sheet["J3"].value = "hww, мм"
    sheet["K3"].value = "sd, мм"
    sheet["L3"].value = "sp, мм"
    sheet["M3"].value = "bd, мм"
    sheet["N3"].value = "Spf, мм"
    sheet["O3"].value = "emax, мм"
    sheet["P3"].value = "emin, мм"
    sheet["Q3"].value = "r, мм"
    sheet["R3"].value = "h, мм"
    sheet["S3"].value = "n"
    sheet["T3"].value = "I∑,мм4"
    sheet["U3"].value = "α"
    sheet["I1"].value = "Ширина заготовки"
    sheet["J1"].value = "Геометрия"
    sheet["V1"].value = "Марка стали"
    ws = book.active
    ws.merge_cells('C1:H1')
    ws.merge_cells('C2:E2')
    ws.merge_cells('F2:H2')
    ws.merge_cells('B1:B3')
    ws.merge_cells('A1:A3')
    ws.merge_cells('I1:I3')
    ws.merge_cells('J1:U2')
    ws.merge_cells('V1:V3')
    font = Font(b = True, size = 11, color = "000000")
    fill1 = PatternFill("solid", fgColor = "FFFF99")
    solid_cell = []
    solid_cell += ["A1", "C1", "C2", "F2", "A1", "B1", "C3", "D3", "E3", "F3", "G3", "H3", "J3", "K3", "L3", "M3", "N3", "O3", "P3", "Q3", "R3", "S3", "T3", "U3", "I1", "J1", "V1"]
    for i in range (0, len(solid_cell)):
        sheet[f'{solid_cell[i]}'].font = font
        sheet[f'{solid_cell[i]}'].fill = fill1
    book.save('formula.xlsx')

def tables():
    book = openpyxl.load_workbook('formula.xlsx')
    book.create_sheet("СТ62-985")
    sheet:worksheet = book.worksheets[1]
    lst = []
    lst = df_to_list(df)
    R = 320
    l = 6
    wx1 = 8.88
    print(lst[0][3])
    col = 0
    row = 3
    q1 = []
    for i in range (0, len(lst)):
        q1 += [round((lst[i][3] * R / (0.125 *l *l * 9.807)), 2)]
    print(q1)
    book.save('formula.xlsx')
   
if __name__ == "__main__":
    writer()
    decor() 
    tables()